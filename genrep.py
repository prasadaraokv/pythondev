# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path_input1 = "d:\\Dev\\Python\\ExcelReport\\.venv\\fn1.xlsx"
path_output1 = "d:\\Dev\\Python\\ExcelReport\\.venv\\o1.xlsx"

# To open the workbook 
# workbook object is created
input1_wb_obj = openpyxl.load_workbook(path_input1)
output1_wb_obj = openpyxl.load_workbook(path_output1)

# Get workbook active sheet object
# from the active attribute
input1_sheet_obj = input1_wb_obj["Sheet1"]
output1_sheet_obj = output1_wb_obj["app"]

# Cell objects also have a row, column, 
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or 
# column integer is 1, not 0.

# Cell object is created by using 
# sheet object's cell() method.
row_index = 4

# Iterate through all rows and columns to find the data range
for row in input1_sheet_obj.iter_rows():
    if row[0].row == 1:
        continue
    
    if row[0].value is not None:  # Check if cell has a value (not empty)

        oname = row[2].value.split(",")
        enum1 = row[3].value #[1:-1]
        aname1 = row[4].value.split(",")
        aname2 = row[6].value.split(",")

        outcellgrpanme = output1_sheet_obj["A" + str(row_index)]
        outcellgrpdsc = output1_sheet_obj["B" + str(row_index)]
        outcellgrpanme.value = row[0].value
        outcellgrpdsc.value = row[0].value

        outcellgrp2fn = output1_sheet_obj["C" + str(row_index)]
        outcellgrp2ln = output1_sheet_obj["D" + str(row_index)]
        outcellgrp2fn.value = oname[0]
        outcellgrp2ln.value = oname[1]

        outcellgrp2enum = output1_sheet_obj["E" + str(row_index)]
        outcellgrp2enum.value = enum1

        outcellgrp3fn = output1_sheet_obj["G" + str(row_index)]
        outcellgrp3ln = output1_sheet_obj["H" + str(row_index)]
        outcellgrp3fn.value = aname1[0]
        outcellgrp3ln.value = aname1[1]

        row_index += 1
    else:
        break

output1_wb_obj.save(path_output1)

